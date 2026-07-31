import openpyxl
import os 
from openpyxl import*
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
path = r"Excel\excel.xlsx"
if not os.path.exists(path):
    workBook = Workbook()
    print(path + " Workbook is created")
    table = Table(displayName="Employee_Table",ref='A1:F8')
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",showFirstColumn=False,showLastColumn=False,showColumnStripes=True,showRowStripes=True)
    workSheet = workBook.active
    workSheet.append(['ID','Name','Age','Language Known','Gender','Comments'])
    gendervalid = DataValidation(type='list',formula1='"Male,Female"',allow_blank=True)
    workSheet.add_data_validation(gendervalid)
    gendervalid.add('E2:E1048576')
    workSheet.add_table(table)
    print(workSheet.max_row)
    print(workSheet.max_column)
else:
    print(path + " path is already exists")
    workBook = load_workbook(path)
    workSheet = workBook.active
    # table = workSheet.tables['Employee_Table']
    # workBook.create_sheet("New Next Sheet")

#  used for current sheet name change
# workSheet.title = 'Student Details'

# workSheet = workBook["New Next Sheet"]
# workSheet = workBook["Student Details"]
# workSheet = workBook["Updated Student Details"]


# how to active the workbook
workBook.active = workBook.index(workSheet)


# using index rename
# workBook[workBook.sheetnames[1]].title = "Updated Student Details"
    
# workSheet['a1'] = "Name"
# workSheet['b1'] = "Year"
# workSheet['c1'] = "Dept"

# workSheet.append(['sethu',1,'cse'])
# workSheet.append(['sethumadavan',1,'IT'])
# workSheet.append(['sethu raman',3,'Bio'])
# workSheet.append(['sethu raja',2,'Mech'])
# workSheet.append(['sethu Kumar',4,'EEE'])
# workSheet.append(['sethu Raj',4,'ECE'])


# workSheet.append(['Naan', 'than da', 'leo'])
# workSheet.append(['leo','das'])

# workSheet.append(['naa','ready','than','vara','vaa','va'])

# workSheet['b2'] = None
# workSheet.delete_rows(4,-1)
# workSheet.delete_rows(5,2)
# workSheet.delete_cols(5)

# table.ref = f'A1:{get_column_letter(workSheet.max_column)}{workSheet.max_row}'
print(workBook.sheetnames)
print(workBook.active)

# print the row in sheet
# for i in workSheet.iter_rows():
#     print(i[0].value,i[1].value,i[2].value)

# for adding comment in the excel
workSheet['F2'].comment = Comment(text='This is a comment',author='Prawin')

workBook.save(path)
workBook.close()
