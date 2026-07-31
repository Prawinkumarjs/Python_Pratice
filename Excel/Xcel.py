# ==========================================================
# Import Required Modules
# ==========================================================

# Used to check whether the Excel file already exists
import os

# Workbook      -> Create a new Excel workbook
# load_workbook -> Open an existing workbook
from openpyxl import Workbook, load_workbook

# Used to create Excel Tables
from openpyxl.worksheet.table import Table, TableStyleInfo

# Used to create dropdown lists (Data Validation)
from openpyxl.worksheet.datavalidation import DataValidation

# Converts column numbers to letters
# Example:
# 1 -> A
# 2 -> B
# 6 -> F
from openpyxl.utils import get_column_letter


# ==========================================================
# File Path
# ==========================================================

# Folder : Excel
# File   : excel.xlsx
path = r"Excel\xcel.xlsx"


# ==========================================================
# Create Workbook if it doesn't exist
# ==========================================================

if not os.path.exists(path):

    print("Workbook doesn't exist.")
    print("Creating Workbook...\n")

    # Create a new workbook
    workbook = Workbook()


    # ======================================================
    # FIRST SHEET (Employees)
    # ======================================================

    # Every new workbook already contains one sheet.
    employeeSheet = workbook.active

    # Rename the default sheet
    employeeSheet.title = "Employees"


    # -------------------------
    # Add Heading Row
    # -------------------------
    employeeSheet.append([
        "ID",
        "Name",
        "Age",
        "Language Known",
        "Gender",
        "Comments"
    ])


    # -------------------------
    # Employee Data
    # -------------------------
    employees = [

        [101, "Prawin", 22, "Python", "Male", "Excellent"],
        [102, "Rahul", 23, "Java", "Male", "Good"],
        [103, "Priya", 21, "C++", "Female", "Very Good"],
        [104, "Arun", 24, "Python", "Male", "Average"],
        [105, "Divya", 23, "SQL", "Female", "Excellent"],
        [106, "Karthik", 25, "Java", "Male", "Good"]

    ]


    # Add every employee row into the sheet
    for employee in employees:
        employeeSheet.append(employee)


    # -------------------------
    # Gender Drop Down
    # -------------------------

    genderValidation = DataValidation(

        type="list",
        formula1='"Male,Female"',
        allow_blank=True

    )

    # Add validation to sheet
    employeeSheet.add_data_validation(genderValidation)

    # Apply dropdown from E2 to E1048576
    genderValidation.add("E2:E1048576")


    # -------------------------
    # Employee Table
    # -------------------------

    employeeTable = Table(

        displayName="Employee_Table",

        ref=f"A1:{get_column_letter(employeeSheet.max_column)}{employeeSheet.max_row}"

    )


    employeeStyle = TableStyleInfo(

        name="TableStyleLight9",

        showFirstColumn=False,

        showLastColumn=False,

        showRowStripes=True,

        showColumnStripes=True

    )

    employeeTable.tableStyleInfo = employeeStyle

    employeeSheet.add_table(employeeTable)



    # ======================================================
    # SECOND SHEET (Students)
    # ======================================================

    # Create a new sheet
    studentSheet = workbook.create_sheet(title="Students")


    # -------------------------
    # Student Heading
    # -------------------------

    studentSheet.append([

        "Roll No",
        "Name",
        "Department",
        "Year",
        "CGPA"

    ])


    # -------------------------
    # Student Data
    # -------------------------

    students = [

        [1, "Sethu", "CSE", 1, 8.2],
        [2, "Hari", "ECE", 2, 8.8],
        [3, "Priya", "IT", 3, 9.1],
        [4, "Ram", "EEE", 4, 8.4],
        [5, "Kumar", "MECH", 2, 7.9],
        [6, "Anitha", "AI", 1, 9.3]

    ]


    # Insert student rows
    for student in students:
        studentSheet.append(student)


    # -------------------------
    # Student Table
    # -------------------------

    studentTable = Table(

        displayName="Student_Table",

        ref=f"A1:{get_column_letter(studentSheet.max_column)}{studentSheet.max_row}"

    )


    studentStyle = TableStyleInfo(

        name="TableStyleMedium2",

        showFirstColumn=False,

        showLastColumn=False,

        showRowStripes=True,

        showColumnStripes=True

    )


    studentTable.tableStyleInfo = studentStyle

    studentSheet.add_table(studentTable)


    # ======================================================
    # THIRD SHEET (Marks)
    # ======================================================

    marksSheet = workbook.create_sheet(title="Marks")


    marksSheet.append([

        "Roll No",
        "Name",
        "Maths",
        "Science",
        "English",
        "Total"

    ])


    marks = [

        [1, "Sethu", 90, 85, 88, 263],
        [2, "Hari", 80, 91, 87, 258],
        [3, "Priya", 95, 94, 96, 285],
        [4, "Ram", 70, 75, 78, 223],
        [5, "Kumar", 88, 84, 81, 253]

    ]


    for mark in marks:
        marksSheet.append(mark)


    marksTable = Table(

        displayName="Marks_Table",

        ref=f"A1:{get_column_letter(marksSheet.max_column)}{marksSheet.max_row}"

    )


    marksStyle = TableStyleInfo(

        name="TableStyleMedium4",

        showFirstColumn=False,

        showLastColumn=False,

        showRowStripes=True,

        showColumnStripes=True

    )


    marksTable.tableStyleInfo = marksStyle

    marksSheet.add_table(marksTable)


    # ======================================================
    # Save Workbook
    # ======================================================

    workbook.save(path)

    print("Workbook Created Successfully!")



# ==========================================================
# Workbook Already Exists
# ==========================================================

else:

    print("Workbook Already Exists.")

    # Open existing workbook
    workbook = load_workbook(path)


# ==========================================================
# Print Workbook Information
# ==========================================================

print("\nWorkbook Information")
print("---------------------------")

# Display all sheet names
print("Sheet Names :", workbook.sheetnames)

# Display active sheet name
print("Active Sheet :", workbook.active.title)

# Number of sheets
print("Total Sheets :", len(workbook.sheetnames))


# ==========================================================
# Save Workbook Again
# ==========================================================

workbook.save(path)

print("\nWorkbook Saved Successfully.")